import openpyxl
import datetime
import platform

################################################################################

def get_save_path():
    if(platform.system() == "Darwin"):
        return "/users/twig/Desktop/"
    elif(platform.system() == "Windows"):
       return "C:/Users/ericl/Desktop/"

################################################################################

def get_workbook_and_stocklist(file):
    stock_list = []
    wb = None
    extension = file[file.rfind(".")+1:]
    if(extension == "xlsx"):
        wb = openpyxl.load_workbook(file)
        stock_list = wb.sheetnames
    elif(extension == "txt"):
        f = open(file, "r")
        wb = openpyxl.Workbook()
        for line in f:
            sheet_title = line
            if(sheet_title[-1] == "\n"):
                sheet_title = sheet_title[:-1]
            stock_list.insert(len(stock_list), sheet_title)
            wb.create_sheet(sheet_title, len(stock_list))
        f.close()
        wb.remove_sheet(wb.get_sheet_by_name("Sheet"))
    return wb, stock_list

################################################################################

# checks if the file is open
# return    True    file is closed (can save)
# return    False   fils is open (can not save)
def fileOpen(wb, fileName, path):
    try:
        wb.save(path + "option_analysis_" + str(datetime.date.today()) + ".xlsx")
        return True
    except PermissionError:
        print("ERROR: FILE IS OPEN")
        print("CLOSE FILE AND RUN SCRIPT AGAIN")
        return False

################################################################################

# saves the completed workbook as option_analysis_ and the current date
# return    True    if the save completed successfully
# return    False   if the save didn't go through
def save(wb, path):
    try:
        wb.save(path + "option_analysis_" + str(datetime.date.today()) + ".xlsx")
        return True
    except PermissionError:
        print("ERROR: FILE IS OPEN")
        print("CLOSE FILE AND RUN SCRIPT AGAIN")
        return False
