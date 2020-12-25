# Eric Liu

# TODO

from pprint import pprint
from tqdm import tqdm

# for errors
import ssl
import requests

import argparse
import openpyxl
import time
import quandl
import json

from src.functions import *
from src.data.run_data import run_data

########################################################################################

def main():

    # read and process command line arguments
    parser = argparse.ArgumentParser()
    parser.add_argument("-c", "--config", type=str, nargs=1, metavar="str", 
        default="config/default.json", help="The path to the desired config file. By default " \
        "this takes the value of \"config/default.json\"")
    parser.add_argument("--test", action="store_true",
        help="When present, the program will override the \"tickers\" parameter of the config " \
        "file to only process on AAPL and ZTE. This will also override the \"--tickers\" argument")
    parser.add_argument("--tickers", type=str, metavar="str", default=None,
        help="Define tickers using a file instead of the config file. This will overwrite the " \
        "tickers included in the config file. The file must contain one ticker per line. See " \
        "TODO for an example file")
    parser.add_argument("-o", "--overwrite", action="store_true",
        help="Download/overwrite existing data. This flag must be passed in everytime the " \
        "tickers change or the data save location changes")
    args = parser.parse_args()

    # open and process the config file
    with open(get_path(args.config), "r") as f:
        config = json.load(f)
    
        # ensure that the save location exists and is a directory
        check_directory(config["save_location"])

        # change the paths to absolute
        config["save_location"] = get_path(config["save_location"])

        # change tickers if the "--test" flag is present
        # also overwrite data
        if(args.test):
            config["tickers"] = ["AAPL", "ZTE"]
            args.overwrite = True

    # get the workbook formatted with the input tickers
    wb = get_workbook(config["tickers"])

    # download/overwrite data if requested
    if(args.overwrite):
        run_data(config["tickers"])

    # save the wb
    save(wb, config["save_location"])

    # start = time.time()

    # try:
    #     save_path = get_save_path()

    #     # get the workbook and stock list
    #     # depends on the file format of the template
    #     template_file = "./templates/template 20-06-2020.txt"
    #     wb, stock_list = get_workbook_and_stocklist(template_file)

    #     data = Data(stock_list)
    #     data.retrieve_data(data_provider = "Quandl")
    #     short_term_data = data.get_short_term_data()

    #     # print(data.data)
    #     # print(short_term_data)

    #     print("\ncreating stock sheets...")
    #     for sheet in wb:
    #         stock_sheet = Stock(sheet, short_term_data[sheet.title])
    #         stock_sheet.format()
    #         stock_sheet.fill_data()
    #         stock_sheet.fill_stats()
    #         stock_sheet.fill_graphs()
    #     print("done")
        
    #     long_term_data = data.get_long_term_data()

    #     # print(long_term_data)

    #     temp_sheet = wb.create_sheet("10YR %", 0)
    #     percentage_sheet = Sheet(temp_sheet, long_term_data["percent_change"], stock_list, long_term_data["years"])
    #     percentage_sheet.format()
    #     percentage_sheet.fill()
    #     percentage_sheet.color(-10, 10)

    #     temp_sheet = wb.create_sheet("10YR % STD DEV", 1)
    #     # note: std dev here is std dev of percentage change, not of average value of the month
    #     std_dev_sheet = Sheet(temp_sheet, long_term_data["std_dev"], stock_list, long_term_data["years"])
    #     std_dev_sheet.format()
    #     std_dev_sheet.fill()
    #     # TODO: do some work to figure out optimal numbers for std dev coloring
    #     std_dev_sheet.color(.1, .02)

    #     temp_sheet = wb.create_sheet("10YR FREQ", 2)
    #     freq_sheet = Sheet(temp_sheet, long_term_data["freq"], stock_list, long_term_data["years"])
    #     freq_sheet.format()
    #     freq_sheet.fill()
    #     freq_sheet.color(20, 80)

    # # this doesn't work to stop no wifi errors for some reason
    # # TODO: catch no wifi error
    # except ConnectionError:
    #     print("\nERROR: NO INTERNET")
    # except requests.exceptions.SSLError:
    #     print("\nERROR: SSL Certificate is not valid")
    # except quandl.errors.quandl_error.QuandlError as err:
    #     print(err)
        
    # time_elapsed = time.time() - start
    # print()
    # print("time elapsed: " + str(int(int(time_elapsed) / 60)) + "m " + str(int(time_elapsed) % 60) + "s")

    # save(wb, save_path)

if __name__ == '__main__':
    main()

