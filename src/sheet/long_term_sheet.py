# Eric Liu

# This file contains all logic needed to take processed (long term) data and fill it into 
# individual sheets for each form of long term data.

import datetime
import os

import pandas as pd

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, PatternFill

from src.functions import make_absolute

def add_long_term_sheets(wb, config):
    """
    Add all long term sheets to the wb based on the config file. This adds the "% Change" sheet, 
    "STD" sheet, and "Frequency" sheet.

    :param:     wb          The workbook to add long term sheets to
    :param:     config      The config file. By default "config/default.json"
    """
    # create and add the percent change sheet
    add_long_term_sheet(wb, config, "monthly_price_std", "Normalized Price STD", 0, std_average)

    # create and add the frequency sheet
    add_long_term_sheet(wb, config, "freq_pct_positive", "Freq(+ % Change)", 0, freq_average)

    # create and add the std sheet
    add_long_term_sheet(wb, config, "std_pct", "STD(% Change)", 0, std_average)

    # create and add the percent change sheet
    add_long_term_sheet(wb, config, "percent_changes", "% Change", 0, perc_average)

def get_month_order(current_month):
    """
    This function takes the current month integer (1=Jan, 12=Dec) and returns the correct order of
    months that should appear in the long term data sheets, including the four month average
    column and ticker column. For example:
    1.  get_month_order(4) should return the list ["ticker", "Apr (4)", "May (5)", "Jun (6)", 
        "Jul (7)", "4 Month", "Aug (8)", "Sep (9)", "Oct (10)", "Nov (11)", "Dec (12)", "Jan (1)", 
        "Feb (2)", "Mar (3)"]
    2.  get_month_order(11) should return the list ["ticker", "Nov (11)", "Dec (12)", "Jan (1)", 
        "Feb (2)",  "4 Month", "Mar (3)", "Apr (4)", "May (5)", "Jun (6)", "Jul (7)", "Aug (8)", 
        "Sep (9)", "Oct (10)"]

    :param:     current_month   An integer representing the current month

    :return:    []              As described above
    """
    # two years of months hard coded (so the list has 24 items)
    months = ["Jan (1)", "Feb (2)", "Mar (3)", "Apr (4)", "May (5)", "Jun (6)", "Jul (7)", 
        "Aug (8)", "Sep (9)", "Oct (10)", "Nov (11)", "Dec (12)", "Jan (1)", "Feb (2)", "Mar (3)", 
        "Apr (4)", "May (5)", "Jun (6)", "Jul (7)", "Aug (8)", "Sep (9)", "Oct (10)", "Nov (11)", 
        "Dec (12)"]

    # get 12 months starting from the input current month
    ordered_months = months[current_month-1:current_month+11]

    # add the 4 month column and ticker column
    ordered_months.insert(4, "4 Month")
    ordered_months.insert(0, "Ticker")

    # if the month is Sep (9), then add an empty column after the four month column 
    if(current_month == 9):
        ordered_months.insert(6, "Empty")
    # otherwise, put the empty column after Dec (12)
    else:
        ordered_months.insert(ordered_months.index("Dec (12)") + 1, "Empty")

    # return the months
    return ordered_months

def freq_average(four_month_freq):
    """
    A function that computes the frequency specific average. This is the probability that the ticker
    will only go up in the four months inputed

    :param:     four_month_freq     A series with the historical positive change frequency
    :return:    float               The historical probability that the stock will only go up in 
                                    the input four months
    """
    return ((four_month_freq / 100).product() * 100)

def std_average(four_month_std):
    """
    A function that computes the std specific average. This is just the average

    :param:     four_month_std      A series with the historical % change std
    :return:    float               The average
    """
    return four_month_std.mean()

def perc_average(four_month_perc):
    """
    A function that computes the % change specific average. This is the total percent change over
    the four months

    :param:     four_month_perc     A series with the historical % change
    :return:    float               The total percent change over the four months
    """
    return ((((four_month_perc / 100) + 1).product() - 1) * 100)

def integer_to_letter(i):
    """
    Return the capital letter associated with the input integer (1=A, 2=B, etc.)

    :param:     i       The integer to change to letter

    :return:    str     The capital letter as a string
    """
    return chr(i + 64)

def add_long_term_sheet(wb, config, file_name, sheet_name, sheet_location, avg_function):
    """
    Add a specific long term sheet using the data from the file_name parameter.

    :param:     wb                  The workbook to add the sheet to
    :param:     config              The config file. By default "config/default.json"
    :param:     file_name           The name (no extension) of the file where data for this sheet
                                    can be found. The file_name should also reflect the default
                                    settings in "config/default.json" since the coloring looks for
                                    the key "{file_name}_low_high" in the config file for how to 
                                    color
    :param:     sheet_name          The name of the sheet to insert
    :param:     sheet_location      Where in the workbook to put the sheet
    :param:     avg_function        The function used to generate the "4 Month column". This 
                                    function should take in a series with four elements and return
                                    a number (usually float)
    """
    # get the current year and month
    current_year = datetime.date.today().year
    current_month = datetime.date.today().month

    # create the sheet at the correct location
    sheet = wb.create_sheet(sheet_name, sheet_location)

    # get the data to fill into the sheet
    
    data_path = os.path.join(make_absolute(config["data_path"]), config["processed_folder"], 
        file_name + ".csv")
    data = pd.read_csv(data_path)

    # get the correct order of months
    month_order = get_month_order(current_month)
    four_month = None
    if(current_month >= 10):
        # if the month is Oct, Nov, or Dec, then there will be an empty column placed after Dec,
        # which will cause issues with the averaging functions. so remove the empty column
        four_month = month_order.copy()[1:6]
        four_month.remove("Empty")
    else:
        four_month = month_order[1:5]

    # compute the four month average based on the specific average function
    data["4 Month"] = data[four_month].apply(avg_function, axis=1)

    # add the empty column
    data["Empty"] = ""

    # reorder the columns
    data = data[month_order]

    # add all the data to the sheet
    for row in dataframe_to_rows(data, index=False, header=True):
        sheet.append(row)

    # remove the "Empty" column header
    # get the letter column it appears in and set value to ""
    column_letter = integer_to_letter(month_order.index("Empty") + 1)
    sheet[column_letter + "1"] = ""

    # format all column headers
    # don't touch ticker, but set all other headers to center alignment
    for letter_int in range(2, 16):
        sheet[integer_to_letter(letter_int) + "1"].alignment = Alignment(horizontal="center")

    # move all cells down one
    sheet.move_range("A1:O" + str(1 + len(data.index)), rows=1, cols=0)

    # merge super header cells and add the years
    # if the current month is January, then all data comes from the previous year
    if(current_month == 1):
        sheet.merge_cells("B1:N1")
        sheet["B1"] = current_year - 1
    # otherwise...
    else:
        # merge all cells up to but not including the empty column
        sheet.merge_cells("B1:" + integer_to_letter(month_order.index("Empty")) + "1")
        sheet["B1"] = current_year - 1
        sheet["B1"].alignment = Alignment(horizontal="center")

        # merge all cells after the empty column
        start_cell = integer_to_letter(month_order.index("Empty") + 2) + "1"
        end_cell = integer_to_letter((month_order.index("Empty") + 2) + (current_month - 2)) + "1"
        sheet.merge_cells(start_cell + ":" + end_cell)
        sheet[start_cell] = current_year
        sheet[start_cell].alignment = Alignment(horizontal="center")

    # color the sheet

    # get the starting cell (top left) and ending cell (bottom right) for the coloring 
    start_cell = "B3"
    end_cell = "O" + str(len(data.index) + 2)

    # get the low/high color theresholds and the color gradient
    low_high = config[file_name + "_low_high"]
    color_gradient = config["color_gradient"]

    # iterate over every cell
    for row_of_cells in sheet[start_cell + ":" + end_cell]:
        for cell in row_of_cells:
            # color the cell based on the value of the cell, ignoring when the cell is empty
            if(cell.value != ""): 
                cell.fill = get_color(low_high[0], low_high[1], float(cell.value), color_gradient)

def get_color(low, high, value, gradient):
    """
    Get the color corresponding to the input value, the low/high values, and the color gradient.
    Note that low can be greater than high, and the color gradient will be flipped accordingly

    :param:     low         The low value (corresponds to red)
    :param:     high        The high value (corresponds to green)
    :param:     value       The value to color
    :param:     gradient    The color gradient to use. By default it is a red (low) to green (high)
                            gradient defined in "config/default.json" with 20 color steps

    :return:    PatternFill     An openpyxl class used to set the background color of a cell. I'm 
                                not sure if this is the best way to do it, but it's the way I did 
                                it before
    """
    # reverse the color gradient if low/high requires it
    if(low > high):
        low, high = (high, low)
        colors = list(reversed(gradient))
    else:
        colors = gradient

    # find the total range between high and low
    r = high - low

    # get the number of steps in the color gradient and the size of each step
    num_steps = len(colors)
    step_size = r / num_steps

    # get the adjusted value (normalized to lowest zero)
    normalized = value - low
    if(normalized < 0):
        normalized = 0
    elif(normalized >= r):
        normalized = r - .00001

    # get the color index
    color_index = int(normalized / step_size)

    # return the color
    return PatternFill(start_color=colors[color_index], end_color=colors[color_index], 
        fill_type="solid")
