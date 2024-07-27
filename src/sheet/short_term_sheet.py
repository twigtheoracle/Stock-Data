# Eric Liu

# This file contains all logic needed to take processed (short term) data and fill it into 
# individual sheets for each stock.

from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment

import pandas as pd

import datetime
import openpyxl

def add_short_term_sheet(ticker, sheet, data, metadata, config):
    """
    This function exists as an easy function to call to create all individual short term sheets

    :param:     ticker      The name of the stock
    :param:     sheet       The sheet to add short term data into
    :param:     data        The short term data corresponding to the stock. This is the contents
                            of the file (default location) "data/processed/{ticker}.csv"
    :param:     metadata    The metadata for the stock. This is one row of the file (default
                            location) "data/processed/metadata.csv"
    :param:     config      The configuration file. Default location is "config/default.json"
    """
    # add all the short term data to the sheet
    for row in dataframe_to_rows(data, index=False, header=True):
        sheet.append(row)

    # format the various columns of short term data in the sheet
    format_short_term(sheet)
    
    # format the various statistics sections
    format_statistics(sheet)

    # add the various statistics to the sheet
    # "Adj_Close" statistics
    sheet["H2"] = metadata["n"].values[0]
    sheet["H3"] = metadata["mean"].values[0]
    sheet["H4"] = metadata["20 Day STD"].values[0]
    sheet["H5"] = metadata["40 Day STD"].values[0]
    sheet["H6"] = metadata["60 Day STD"].values[0]

    # IV Statistics
    # sheet["H10"] = metadata["next_earnings_day"].values[0]
    # sheet["H11"] = metadata["trading_days"].values[0]
    # sheet["H12"] = metadata["calendar_days"].values[0]
    # sheet["H13"] = metadata["crush_rate"].values[0]
    sheet["H10"] = "N/A"
    sheet["H11"] = "N/A"
    sheet["H12"] = "N/A"
    sheet["H13"] = "N/A"

    # add the bin averages/counts to the sheet
    add_bins(sheet, data["Adj_Close"], config["num_bins"])

    # add the charts
    add_charts(sheet, config["num_bins"])
    
def format_short_term(sheet):
    """
    Format the columns "A" through "E" which contain the short term raw data

    :param:     sheet       The sheet to format
    """
    # "Date" column
    # just set the width
    date_col = sheet.column_dimensions["A"]
    date_col.width = 11

    # "Adj_Close" column
    # set width, change the number format
    adj_close_col = sheet.column_dimensions["B"]
    adj_close_col.width = 10
    for cell in sheet["B"]:
        cell.number_format = "0.00"

    # IV columns
    # set width for all
    for col_letter in letter_range("C", "E"):
        iv_col = sheet.column_dimensions[col_letter]
        iv_col.width = 11

    # set the alignment of all headers to center
    for col_letter in letter_range("A", "E"):
        sheet[col_letter + "1"].alignment = Alignment(horizontal="center")

def format_statistics(sheet):
    """
    Format the cells that contain Adj_Close statistics, IV statistics, and histogram data

    :param:     sheet       The sheet to format
    """
    # set the width of the columns
    name_col = sheet.column_dimensions["G"]
    name_col.width = 18
    value_col = sheet.column_dimensions["H"]
    value_col.width = 10

    # "Price Statistics" header
    sheet.merge_cells("G1:H1")
    sheet["G1"] = "Adj_Close Statistics"
    sheet["G1"].alignment = Alignment(horizontal="center")

    # add in the names of the various price statistics
    sheet["G2"] = "n"
    sheet["G3"] = "mean"
    sheet["G4"] = "20 Day STD"
    sheet["G5"] = "40 Day STD"
    sheet["G6"] = "60 Day STD"

    # set the number format of price statistics to two decimal places
    for row_number in range(3, 7):
        sheet["H" + str(row_number)].number_format = "0.00"

    # "IV Statistics" header and date
    sheet.merge_cells("G8:H8")
    sheet["G8"] = "IV Statistics"
    sheet["G8"].alignment = Alignment(horizontal="center")
    sheet.merge_cells("G9:H9")
    sheet["G9"] = f"Data from {str(datetime.date.today())}"
    sheet["G9"].alignment = Alignment(horizontal="center")

    # add in the names of the various iv statistics
    sheet["G10"] = "Next Earnings Day"
    sheet["G11"] = "Trading Days"
    sheet["G12"] = "Calendar Days"
    sheet["G13"] = "Earnings Crush Rate"

    # "Histogram Data" header
    sheet.merge_cells("G15:H15")
    sheet["G15"] = "Histogram Data"
    sheet["G15"].alignment = Alignment(horizontal="center")

    # "Histogram Data" column names
    sheet["G16"] = "Bin Average"
    sheet["G16"].alignment = Alignment(horizontal="center")
    sheet["H16"] = "Count"
    sheet["H16"].alignment = Alignment(horizontal="center")

def add_bins(sheet, data, bins):
    """
    Compute bins and counts and put them into the sheet 

    :param:     sheet       The sheet to add bins/counts into
    :param:     data        The data to gererate bins/counts from
    :param:     bins        The number of bins as defined in the config file
    """
    # get the bins and counts
    bins_counts = data.value_counts(bins=bins).to_frame().reset_index()

    # get the average value of each bin and sort
    bins_counts["bin_average"] = bins_counts["index"].apply(lambda x: x.mid)
    bins_counts = bins_counts.sort_values("bin_average", ascending=True)

    # add the bin averages and counts to the sheet
    for i in range(0, bins):
        sheet["G" + str(i+17)] = bins_counts["bin_average"].values[i]
        try:
            sheet["H" + str(i+17)] = bins_counts["Adj_Close"].values[i]
        except:
            pass
        try:
            sheet["H" + str(i+17)] = bins_counts["count"].values[i]
        except:
            pass

def add_charts(sheet, bins):
    """
    Add the three charts of Adj_Close histogram, Adj_Close line chart, and IV line chart

    :param:     sheet       The sheet to add charts into
    :param:     bins        The number of bins as defined in the config file
    """
    # histogram of adj_close prices
    histogram = openpyxl.chart.BarChart()
    histogram.shape = 4
    histogram.type = "col"
    histogram.style = 10
    histogram.title = sheet.title + " Adj_Close Price Histogram"
    histogram.x_axis.title = "Bin Average Adj_Close Price"
    histogram.y_axis.title = "Count"
    histogram.legend = None
    histogram_categories = openpyxl.chart.Reference(sheet, min_col=7, max_col=7, min_row=17, 
        max_row=(17 + bins - 1))
    histogram_data = openpyxl.chart.Reference(sheet, min_col=8, max_col=8, min_row=17, 
        max_row=(17 + bins - 1))
    histogram.add_data(histogram_data, titles_from_data=False)
    histogram.set_categories(histogram_categories)
    sheet.add_chart(histogram, "J2")

    # line chart of adj_close prices
    line_chart = openpyxl.chart.LineChart()
    line_chart.style = 12
    line_chart.title = sheet.title + " Adj_Close Line Chart"
    line_chart.x_axis_title = "Date"
    line_chart.y_axis_title = "Adj_Close Price"
    line_chart.legend = None
    line_data = openpyxl.chart.Reference(sheet, min_col=2, max_col=2, min_row=2, max_row=61)
    line_categories = openpyxl.chart.Reference(sheet, min_col=1, max_col=1, min_row=2, max_row=61)
    line_chart.add_data(line_data)
    line_chart.set_categories(line_categories)
    # style the line
    style = line_chart.series[0]
    style.graphicalProperties.line.width = 25000
    style.smooth = True
    sheet.add_chart(line_chart, "J17")

    # line chart of iv data
    iv_chart = openpyxl.chart.LineChart()
    iv_chart.style = 12
    iv_chart.title = sheet.title + " IV Data Line Chart"
    iv_chart.x_axis_title = "Date"
    iv_chart.y_axis_title = "Adj_Close Price"
    iv_data = openpyxl.chart.Reference(sheet, min_col=3, max_col=5, min_row=1, max_row=61)
    iv_categories = openpyxl.chart.Reference(sheet, min_col=1, max_col=1, min_row=2, max_row=61)
    iv_chart.add_data(iv_data, titles_from_data=True)
    iv_chart.set_categories(iv_categories)
    # style the lines
    colors = ["00FF00", "FF0000", "0000FF"]
    for series_num in range(0, 3):
        style = iv_chart.series[series_num]
        style.graphicalProperties.line.width = 10000
        style.graphicalProperties.line.solidFill = colors[series_num]
        style.smooth = True
    sheet.add_chart(iv_chart, "J32")

def letter_range(c1, c2):
    """
    Get the letter range between the two input characters inclusive. I'm not sure what will happen 
    if you don't input two capital letters in reverse order, like ("E", "A") so don't do that.
    Code originally from https://stackoverflow.com/questions/7001144/range-over-character-in-python

    :param:     c1      The first character
    :param:     c2      The second character
    """
    """
    >>> for letter in letter_range("A", "E"):
    >>>     print(letter)
    >>> A
    >>> B
    >>> C
    >>> D
    >>> E
    """
    for c in range(ord(c1), ord(c2) + 1):
        yield chr(c)
