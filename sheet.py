from openpyxl.styles import Color, PatternFill

import openpyxl
import datetime
import string
import time

import api_key as key

# returns an uppcase representation of the number
# 1 --> A, 2 --> B, etc.
def number_to_letter(n):
    return string.ascii_uppercase[n - 1]

# class for all the long term sheets
class Sheet():
    # initializes the sheet with basic data
    def __init__(self, sheet, data, stock_list):
        self.sheet = sheet
        self.data = data
        self.stock_list = stock_list

        self.month_index = ["Jan (1)", "Feb (2)", "Mar (3)", "Apr (4)", "May (5)", "Jun (6)", "Jul (7)", "Aug (8)", "Sep (9)", "Oct (10)", "Nov (11)", "Dec (12)"]
        self.color_gradient = [PatternFill(start_color='E4001A', end_color='E4001A', fill_type='solid'), 
            PatternFill(start_color='E30201', end_color='E30201', fill_type='solid'), 
            PatternFill(start_color='E31c02', end_color='E31c02', fill_type='solid'), 
            PatternFill(start_color='E33602', end_color='E33602', fill_type='solid'), 
            PatternFill(start_color='E25003', end_color='E25003', fill_type='solid'), 
            PatternFill(start_color='E26904', end_color='E26904', fill_type='solid'), 
            PatternFill(start_color='E28304', end_color='E28304', fill_type='solid'), 
            PatternFill(start_color='E19C05', end_color='E19C05', fill_type='solid'), 
            PatternFill(start_color='E1B506', end_color='E1B506', fill_type='solid'), 
            PatternFill(start_color='E1CD06', end_color='E1CD06', fill_type='solid'), 
            PatternFill(start_color='DBE007', end_color='DBE007', fill_type='solid'), 
            PatternFill(start_color='C2E008', end_color='C2E008', fill_type='solid'), 
            PatternFill(start_color='A9E008', end_color='A9E008', fill_type='solid'), 
            PatternFill(start_color='91DF09', end_color='91DF09', fill_type='solid'), 
            PatternFill(start_color='79DF09', end_color='79DF09', fill_type='solid'), 
            PatternFill(start_color='60DF0A', end_color='60DF0A', fill_type='solid'), 
            PatternFill(start_color='49DE0B', end_color='49DE0B', fill_type='solid'), 
            PatternFill(start_color='31DE0B', end_color='31DE0B', fill_type='solid'), 
            PatternFill(start_color='1ADE0C', end_color='1ADE0C', fill_type='solid'), 
            PatternFill(start_color='0CDE17', end_color='0CDE17', fill_type='solid')]

    # formats the sheet
    def format(self):
        this_month = int(str(datetime.datetime.now())[5:7])

        for row in range(0, len(self.data)):
            # puts stocks in the A column
            self.sheet["A" + str(row + 3)] = self.stock_list[row]

        # puts the months in the second row
        offset = 0
        index = 2
        for month in range(this_month - 1, this_month + 11):
            self.sheet[number_to_letter(index + offset) + "2"] = self.month_index[month % 12]
            self.sheet[number_to_letter(index + offset) + "2"].alignment = openpyxl.styles.Alignment(horizontal='center')
            index += 1
            if(month == 11):
                offset = 1

        # merges and puts the year(s) into the top row
        month_offset = 12 - this_month + 1
        first_range = "B1:" + number_to_letter(1 + month_offset) + "1"
        second_range = number_to_letter(1 + month_offset + 2) + "1:" + number_to_letter(2 + month_offset + (12 - month_offset)) + "1"
        self.sheet.merge_cells(first_range)
        self.sheet.merge_cells(second_range)
        self.sheet["B1"] = int(str(datetime.datetime.now())[:4])
        self.sheet["B1"].alignment = openpyxl.styles.Alignment(horizontal='center')
        self.sheet[second_range[:2]] = int(str(datetime.datetime.now())[:4]) + 1
        self.sheet[second_range[:2]].alignment = openpyxl.styles.Alignment(horizontal='center')

    # fills the sheet with data
    def fill(self):
        for row in range(0, len(self.stock_list)):
            for column in range(2, 15):
                self.sheet[number_to_letter(column) + str(row + 3)] = self.data[self.stock_list[row]][column - 2]

    # colors the sheet based on a low and high value
    def color(self, red, green):
        low = None
        high = None
        if(red < green):
            low = red
            high = green
        else:
            low = green
            high = red
        bin_size = (high - low) / len(self.color_gradient)

        for row in range(3, 3 + len(self.stock_list)):   
            for column in range(2, 15):
                cell = number_to_letter(column) + str(row)
                if(self.sheet[cell].value != None):
                    index = int((self.sheet[cell].value - low) / bin_size)
                    if(index < 0):
                        index = 0
                    elif(index >= len(self.color_gradient)):
                        index = len(self.color_gradient) - 1
                    self.sheet[cell].fill = self.color_gradient[index]

    # colors the sheet from a low and high percentage value where perecentages are the sheets data compared to the data_list
    def color_percentage(self, red, green, data_list):
        low = None
        high = None
        if(red < green):
            low = red
            high = green
        else:
            low = green
            high = red
        bin_size = (high - low) / len(self.color_gradient)

        for row in range(3, 3 + len(self.stock_list)):  
            stock_name = self.sheet["A" + str(row)].value
            for column in range(2, 15):  
                data_cell = number_to_letter(column) + str(row)
                if(self.sheet[data_cell].value != None):
                    percentage = (self.sheet[data_cell].value / data_list[stock_name][column - 2]) * 100
                    index = int(percentage / bin_size)
                    if(index < 0):
                        index = 0
                    elif(index >= len(self.color_gradient)):
                        index = len(self.color_gradient) - 1
                    self.sheet[data_cell].fill = self.color_gradient[index]



    # # colors each numerical percentage a color based on a pretty gradient from red to green
    # def color(self):
    #     for letter in range(2, 15): # from letter B to N when plugged into the first non __init__ function
    #         for row in range(3, 3 + len(self.stockList)):
    #             cell = self.numberToLetter(letter) + str(row)
    #             value = self.sheet[cell].value

    #             if (value != None):
    #                 value = int(value + 10)
    #                 if (value <= 0):
    #                     value = 0
    #                 elif (value >= 19):
    #                     value = 19
    #                 self.sheet[cell].fill = self.colorGradient[value]